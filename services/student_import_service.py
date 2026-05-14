from typing import Any, Callable, Dict, Optional, Set, Tuple

import openpyxl
from services.student_service import StudentService
from services.import_validators.student_import_validator import StudentImportValidator
from utils.helpers import normalize_classroom_name, normalize_enrollment_number, normalize_text
from utils.importers.excel_student_importer import ExcelStudentImporter
from utils.importers.import_models import ImportResult, ImportError, ImportedRow
from database.repository import student_repo, classroom_repo


ProgressCallback = Optional[Callable[[int, int, int, str, int, int, int], None]]


class StudentImportService:
    """Servicio que orquesta la importación secuencial de estudiantes desde Excel."""

    def __init__(self, student_service: Optional[StudentService] = None):
        self.student_service = student_service if student_service else StudentService()
        self.importer = ExcelStudentImporter()
        self.validator = StudentImportValidator()
        self.classroom_map = self._build_classroom_map()
        self.existing_keys = self._build_existing_keys()

    def validate_structure(self, file_path: str) -> Tuple[bool, Optional[str]]:
        return self.importer.validate_structure(file_path)

    def preview_rows(self, file_path: str, max_rows: int = 15):
        return self.importer.preview_rows(file_path, max_rows=max_rows)

    def import_from_excel(self, file_path: str, progress_callback: ProgressCallback = None) -> ImportResult:
        result = ImportResult(
            total_rows=0,
            imported_count=0,
            omitted_count=0,
            error_count=0,
            errors=[],
            omitted_rows=[],
            success_rows=[]
        )

        total_rows = self._count_data_rows(file_path)
        result.total_rows = total_rows

        processed = 0
        for row_index, row_data in self.importer.iter_rows(file_path):
            processed += 1
            status_message = f"Procesando fila {row_index}"
            self._emit_progress(progress_callback, processed, total_rows, row_index, status_message,
                                 result.imported_count, result.omitted_count, result.error_count)

            valid, validation_message = self.validator.validate_row_data(row_data, row_index)
            student_name = f"{row_data.get('nombre', '')} {row_data.get('apellido', '')}".strip()

            if not valid:
                result.error_count += 1
                result.errors.append(ImportError(
                    row_number=row_index,
                    student_name=student_name,
                    field='fila',
                    error_type='Validación',
                    message=validation_message or 'Datos inválidos'
                ))
                continue

            classroom_id, classroom_label = self._resolve_classroom(row_data)
            if classroom_id is None:
                result.error_count += 1
                result.errors.append(ImportError(
                    row_number=row_index,
                    student_name=student_name,
                    field='aula',
                    error_type='Aula no encontrada',
                    message='No se pudo mapear Curso/Aula/Turno a un aula existente'
                ))
                continue

            duplicate_key = self._build_duplicate_key(row_data, classroom_id)
            if duplicate_key in self.existing_keys:
                result.omitted_count += 1
                result.omitted_rows.append(ImportedRow(
                    row_number=row_index,
                    first_name=self._normalize_text(row_data.get('nombre')),
                    last_name=self._normalize_text(row_data.get('apellido')),
                    classroom_name=classroom_label,
                    academic_year=self._parse_year(row_data.get('año escolar')),
                    status='omitted',
                    notes='Estudiante ya existe en el curso/aula actual',
                    raw_data=row_data
                ))
                continue

            student_data = self._build_student_data(row_data, classroom_id)
            try:
                self.student_service.create_student(student_data)
                result.imported_count += 1
                result.success_rows.append(ImportedRow(
                    row_number=row_index,
                    first_name=student_data.get('first_name', ''),
                    last_name=student_data.get('last_name', ''),
                    classroom_name=classroom_label,
                    academic_year=self._parse_year(row_data.get('año escolar')),
                    status='imported',
                    notes='Importado correctamente',
                    raw_data=row_data
                ))
                self.existing_keys.add(duplicate_key)
            except Exception as exc:
                result.error_count += 1
                result.errors.append(ImportError(
                    row_number=row_index,
                    student_name=student_name,
                    field='servicio',
                    error_type='Base de datos',
                    message=str(exc)
                ))

        self._emit_progress(progress_callback, total_rows, total_rows, total_rows, 'Importación completada',
                             result.imported_count, result.omitted_count, result.error_count)
        return result

    def _count_data_rows(self, file_path: str) -> int:
        rows = 0
        workbook = None
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            worksheet = workbook.active
            for _ in worksheet.iter_rows(min_row=2, values_only=True):
                rows += 1
        except Exception:
            rows = 0
        finally:
            if workbook is not None:
                workbook.close()
        return rows

    def _build_classroom_map(self):
        classrooms = classroom_repo.get_with_details()
        mapping = {}
        for classroom in classrooms:
            grade_name = normalize_classroom_name(classroom.get('grade_name', ''))
            room_name = normalize_classroom_name(classroom.get('name', ''))
            shift = normalize_text(classroom.get('shift', ''))
            mapping[(grade_name, room_name, shift)] = classroom['id']
        return mapping

    def _build_existing_keys(self) -> Set[Tuple[Any, ...]]:
        existing = set()
        for student in student_repo.get_all():
            enrollment_number = normalize_enrollment_number(student.get('enrollment_number'))
            if enrollment_number:
                existing.add(('enrollment_number', enrollment_number))

            if student.get('classroom_id'):
                existing.add((
                    'student_classroom',
                    normalize_text(student.get('first_name')),
                    normalize_text(student.get('last_name')),
                    student['classroom_id']
                ))
        return existing

    def _resolve_classroom(self, row_data: Dict[str, Any]) -> Tuple[Optional[int], Optional[str]]:
        curso = normalize_classroom_name(row_data.get('curso'))
        aula = normalize_classroom_name(row_data.get('aula'))
        turno = normalize_text(row_data.get('turno'))

        if not curso or not aula:
            return None, None

        classroom_id = self.classroom_map.get((curso, aula, turno))
        if classroom_id is not None:
            return classroom_id, f"{row_data.get('curso', '')} / {row_data.get('aula', '')} / {row_data.get('turno', '')}"

        for (grade_name, room_name, shift), cid in self.classroom_map.items():
            if grade_name == curso and room_name == aula:
                return cid, f"{row_data.get('curso', '')} / {row_data.get('aula', '')} / {shift}"

        return None, None

    def _build_duplicate_key(self, row_data: Dict[str, Any], classroom_id: int) -> Tuple[Any, ...]:
        enrollment_number = normalize_enrollment_number(row_data.get('matrícula'))
        if enrollment_number:
            return ('enrollment_number', enrollment_number)
        return (
            'student_classroom',
            normalize_text(row_data.get('nombre')),
            normalize_text(row_data.get('apellido')),
            classroom_id
        )

    def _build_student_data(self, row_data: Dict[str, Any], classroom_id: int) -> Dict[str, Any]:
        student_data = {
            'first_name': normalize_text(row_data.get('nombre')),
            'last_name': normalize_text(row_data.get('apellido')),
            'nivel': normalize_text(row_data.get('nivel')),
            'classroom_id': classroom_id,
            'enrollment_number': normalize_enrollment_number(row_data.get('matrícula')),
            'tutor_name': normalize_text(row_data.get('tutor')),
            'origin_center': normalize_text(row_data.get('origen')),
            'enrollment_status': 'activo'
        }
        return {k: v for k, v in student_data.items() if v or k in ['classroom_id', 'enrollment_number', 'enrollment_status']}

    def _parse_year(self, value: Any) -> Optional[int]:
        try:
            return int(value)
        except Exception:
            return None

    def _normalize_text(self, value: Any) -> str:
        if value is None:
            return ''
        return str(value).strip()

    def _emit_progress(self, callback: ProgressCallback, processed: int, total: int, row_number: int, message: str,
                       imported_count: int = 0, omitted_count: int = 0, error_count: int = 0):
        if callback:
            callback(processed, total, row_number, message, imported_count, omitted_count, error_count)
