"""
Excel Importer - Módulo para importación de calificaciones desde Excel
Permite a los profesores importar notas usando una plantilla estándar
"""

from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from services.academic_service import academic_service
from database.repository import student_repo, subject_repo


@dataclass
class ImportError:
    """Representa un error durante la importación"""
    row_number: int
    student_name: str
    subject: str
    error_type: str
    message: str


@dataclass
class ImportedGrade:
    """Representa una calificación importada exitosamente"""
    row_number: int
    student_name: str
    subject: str
    score: float
    recovery_score: Optional[float]


@dataclass
class ImportResult:
    """Resultado del proceso de importación"""
    total_rows: int
    imported_count: int
    error_count: int
    errors: List[ImportError]
    success_details: List[ImportedGrade]
    
    def get_summary(self) -> str:
        """Resumen del resultado"""
        return f"Importación completada: {self.imported_count} exitosas, {self.error_count} errores de {self.total_rows} filas"


class ExcelValidator:
    """Validaciones de estructura y datos del Excel"""
    
    REQUIRED_COLUMNS = ['Nombre', 'Apellido', 'Materia', 'Trimestre', 'Nota', 'Nota_Recuperacion']
    
    @staticmethod
    def validate_columns(worksheet) -> Tuple[bool, List[str]]:
        """Valida que las columnas requeridas existan"""
        header_row = worksheet[1]
        column_names = [cell.value for cell in header_row]
        
        missing_columns = []
        for required_col in ExcelValidator.REQUIRED_COLUMNS:
            if required_col not in column_names:
                missing_columns.append(required_col)
        
        return len(missing_columns) == 0, missing_columns
    
    @staticmethod
    def validate_grade_data(row_data: Dict[str, Any], row_number: int) -> Tuple[bool, Optional[str]]:
        """Valida los datos de una fila"""
        # Validar campos obligatorios no vacíos
        if not row_data.get('Nombre') or not str(row_data['Nombre']).strip():
            return False, "El nombre del estudiante es obligatorio"
        
        if not row_data.get('Apellido') or not str(row_data['Apellido']).strip():
            return False, "El apellido del estudiante es obligatorio"
        
        if not row_data.get('Materia') or not str(row_data['Materia']).strip():
            return False, "El nombre de la materia es obligatorio"
        
        # Validar trimestre
        trimester = row_data.get('Trimestre')
        if trimester is None:
            return False, "El trimestre es obligatorio"
        
        try:
            trimester = int(trimester)
            if trimester not in [1, 2, 3]:
                return False, "El trimestre debe ser 1, 2 o 3"
        except (ValueError, TypeError):
            return False, "El trimestre debe ser un número válido (1, 2 o 3)"
        
        # Validar nota
        score = row_data.get('Nota')
        if score is None:
            return False, "La nota es obligatoria"
        
        try:
            score = float(score)
            if not (0 <= score <= 10):
                return False, "La nota debe estar entre 0 y 10"
        except (ValueError, TypeError):
            return False, "La nota debe ser un número válido"
        
        # Validar nota de recuperación (opcional)
        recovery_score = row_data.get('Nota_Recuperacion')
        if recovery_score is not None and recovery_score != '':
            try:
                recovery_score = float(recovery_score)
                if not (0 <= recovery_score <= 10):
                    return False, "La nota de recuperación debe estar entre 0 y 10"
            except (ValueError, TypeError):
                return False, "La nota de recuperación debe ser un número válido"
        
        return True, None


class DataMapper:
    """Mapeo de nombres a IDs de la base de datos"""
    
    def __init__(self):
        # Cargar caché de estudiantes y materias
        self._students_cache = None
        self._subjects_cache = None
    
    def _get_students_cache(self) -> Dict[str, int]:
        """Obtiene caché de estudiantes (nombre_completo -> id)"""
        if self._students_cache is None:
            students = student_repo.get_all()
            self._students_cache = {}
            for student in students:
                full_name = f"{student['first_name']} {student['last_name']}".strip().lower()
                self._students_cache[full_name] = student['id']
        return self._students_cache
    
    def _get_subjects_cache(self) -> Dict[str, int]:
        """Obtiene caché de materias (nombre -> id)"""
        if self._subjects_cache is None:
            subjects = subject_repo.get_all()
            self._subjects_cache = {}
            for subject in subjects:
                self._subjects_cache[subject['name'].strip().lower()] = subject['id']
        return self._subjects_cache
    
    def map_student_to_id(self, first_name: str, last_name: str) -> Optional[int]:
        """Busca estudiante por nombre y apellido"""
        full_name = f"{first_name.strip()} {last_name.strip()}".lower()
        students_cache = self._get_students_cache()
        return students_cache.get(full_name)
    
    def map_subject_to_id(self, subject_name: str) -> Optional[int]:
        """Busca materia por nombre"""
        subject_clean = subject_name.strip().lower()
        subjects_cache = self._get_subjects_cache()
        return subjects_cache.get(subject_clean)


class ExcelImporter:
    """Orquestador principal del proceso de importación"""
    
    def __init__(self):
        self.validator = ExcelValidator()
        self.mapper = DataMapper()
    
    def validate_excel_structure(self, file_path: str) -> Tuple[bool, Optional[str]]:
        """Valida la estructura básica del archivo Excel"""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            worksheet = workbook.active
            
            # Validar columnas
            columns_valid, missing_columns = self.validator.validate_columns(worksheet)
            
            workbook.close()
            
            if not columns_valid:
                error_msg = f"Columnas faltantes: {', '.join(missing_columns)}"
                return False, error_msg
            
            return True, None
            
        except Exception as e:
            return False, f"Error al leer el archivo Excel: {str(e)}"
    
    def import_grades_from_excel(self, file_path: str, academic_year: int, teacher_id: int) -> ImportResult:
        """Importa calificaciones desde archivo Excel"""
        
        result = ImportResult(
            total_rows=0,
            imported_count=0,
            error_count=0,
            errors=[],
            success_details=[]
        )
        
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            worksheet = workbook.active
            
            # Validar estructura
            columns_valid, missing_columns = self.validator.validate_columns(worksheet)
            if not columns_valid:
                result.errors.append(ImportError(
                    row_number=0,
                    student_name="",
                    subject="",
                    error_type="Estructura",
                    message=f"Columnas faltantes: {', '.join(missing_columns)}"
                ))
                return result
            
            # Obtener índices de columnas
            header_row = worksheet[1]
            column_indices = {}
            for idx, cell in enumerate(header_row, 1):
                if cell.value:
                    column_indices[cell.value] = idx
            
            # Procesar filas de datos (empezando desde la fila 2)
            data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            result.total_rows = len(data_rows)
            
            for row_idx, row in enumerate(data_rows, start=2):  # row_idx es el número real de fila
                
                # Extraer datos de la fila
                row_data = {}
                for col_name, col_idx in column_indices.items():
                    if col_idx <= len(row):
                        row_data[col_name] = row[col_idx - 1]  # -1 porque row es 0-indexed
                
                student_name = f"{row_data.get('Nombre', '')} {row_data.get('Apellido', '')}".strip()
                subject_name = str(row_data.get('Materia', '')).strip()
                
                # Validar datos de la fila
                is_valid, error_msg = self.validator.validate_grade_data(row_data, row_idx)
                if not is_valid:
                    result.errors.append(ImportError(
                        row_number=row_idx,
                        student_name=student_name,
                        subject=subject_name,
                        error_type="Datos",
                        message=error_msg
                    ))
                    result.error_count += 1
                    continue
                
                # Mapear a IDs
                student_id = self.mapper.map_student_to_id(
                    str(row_data['Nombre']), 
                    str(row_data['Apellido'])
                )
                
                if not student_id:
                    result.errors.append(ImportError(
                        row_number=row_idx,
                        student_name=student_name,
                        subject=subject_name,
                        error_type="Estudiante",
                        message="Estudiante no encontrado en el sistema"
                    ))
                    result.error_count += 1
                    continue
                
                subject_id = self.mapper.map_subject_to_id(subject_name)
                if not subject_id:
                    result.errors.append(ImportError(
                        row_number=row_idx,
                        student_name=student_name,
                        subject=subject_name,
                        error_type="Materia",
                        message="Materia no encontrada en el sistema"
                    ))
                    result.error_count += 1
                    continue
                
                # Preparar datos para la base de datos
                score_data = {
                    'student_id': student_id,
                    'subject_id': subject_id,
                    'teacher_id': teacher_id,
                    'trimester': int(row_data['Trimestre']),
                    'score': float(row_data['Nota']),
                    'academic_year': academic_year
                }
                
                # Agregar nota de recuperación si existe
                recovery_score = row_data.get('Nota_Recuperacion')
                if recovery_score is not None and recovery_score != '':
                    score_data['recovery_score'] = float(recovery_score)
                
                # Registrar calificación
                try:
                    score_id = academic_service.add_score(score_data)
                    
                    # Registrar éxito
                    result.success_details.append(ImportedGrade(
                        row_number=row_idx,
                        student_name=student_name,
                        subject=subject_name,
                        score=float(row_data['Nota']),
                        recovery_score=float(recovery_score) if recovery_score and recovery_score != '' else None
                    ))
                    result.imported_count += 1
                    
                except Exception as e:
                    result.errors.append(ImportError(
                        row_number=row_idx,
                        student_name=student_name,
                        subject=subject_name,
                        error_type="Base de datos",
                        message=f"Error al guardar calificación: {str(e)}"
                    ))
                    result.error_count += 1
            
            workbook.close()
            return result
            
        except Exception as e:
            result.errors.append(ImportError(
                row_number=0,
                student_name="",
                subject="",
                error_type="Sistema",
                message=f"Error general en importación: {str(e)}"
            ))
            return result
    
    def generate_template(self, output_path: str) -> bool:
        """Genera una plantilla Excel vacía con las columnas correctas"""
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Importación de Notas"
            
            # Configurar encabezados
            headers = ['Nombre', 'Apellido', 'Materia', 'Trimestre', 'Nota', 'Nota_Recuperacion']
            
            # Estilos para encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Escribir encabezados
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Ejemplos de datos
            example_data = [
                ['Juan', 'Pérez', 'Matemáticas', 1, 7.5, ''],
                ['María', 'García', 'Lenguaje', 1, 4.0, 6.0],
                ['Carlos', 'López', 'Ciencias', 1, 8.2, ''],
            ]
            
            for row_idx, row_data in enumerate(example_data, start=2):
                for col_idx, value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Ajustar ancho de columnas
            column_widths = [15, 15, 20, 10, 10, 18]
            for col_idx, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
            
            # Agregar hoja de instrucciones
            instruction_sheet = workbook.create_sheet("Instrucciones")
            
            instructions = [
                "INSTRUCCIONES PARA IMPORTAR NOTAS",
                "",
                "1. Complete los datos siguiendo el formato de la hoja 'Importación de Notas'",
                "2. Columnas obligatorias:",
                "   - Nombre: Nombre del estudiante",
                "   - Apellido: Apellido del estudiante", 
                "   - Materia: Nombre exacto de la materia",
                "   - Trimestre: Número 1, 2 o 3",
                "   - Nota: Calificación (0 a 10)",
                "   - Nota_Recuperacion: Opcional, solo si aplica",
                "",
                "3. Los nombres de estudiantes y materias deben coincidir exactamente",
                "4. Las notas deben estar en el rango de 0 a 10",
                "5. El trimestre debe ser 1, 2 o 3",
                "",
                "6. No modifique los encabezados",
                "7. Guarde el archivo y use la opción de importación en GES",
            ]
            
            for row_idx, instruction in enumerate(instructions, 1):
                instruction_sheet.cell(row=row_idx, column=1, value=instruction)
            
            # Guardar archivo
            workbook.save(output_path)
            return True
            
        except Exception as e:
            print(f"Error al generar plantilla: {str(e)}")
            return False


# Instancia global del importador
excel_importer = ExcelImporter()
