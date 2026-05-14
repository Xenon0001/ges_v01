from typing import Any, Dict, Iterator, List, Optional, Tuple
from pathlib import Path
from datetime import date

import openpyxl
from openpyxl import Workbook

from .import_models import PreviewRow
from services.import_validators.student_import_validator import StudentImportValidator


class ExcelStudentImporter:
    """Importador de estudiantes desde Excel usando openpyxl en modo read_only."""

    EXPECTED_COLUMNS = [
        'nombre',
        'apellido',
        'curso',
        'aula',
        'turno',
        'tutor',
        'matrícula',
        'género',
        'fechanacimiento',
        'teléfono',
        'año escolar'
    ]

    COLUMN_ALIASES = {
        'nombre': 'nombre',
        'apellido': 'apellido',
        'curso': 'curso',
        'aula': 'aula',
        'turno': 'turno',
        'tutor': 'tutor',
        'matrícula': 'matrícula',
        'matricula': 'matrícula',
        'género': 'género',
        'genero': 'género',
        'fechanacimiento': 'fechanacimiento',
        'fecha nacimiento': 'fechanacimiento',
        'teléfono': 'teléfono',
        'telefono': 'teléfono',
        'año escolar': 'año escolar',
        'ano escolar': 'año escolar',
        'año': 'año escolar',
        'year': 'año escolar'
    }

    def __init__(self):
        self.validator = StudentImportValidator()

    def _normalize_header(self, header_row: Tuple[Any, ...]) -> Dict[str, int]:
        normalized = {}
        for index, raw_value in enumerate(header_row, start=1):
            if raw_value is None:
                continue
            key = str(raw_value).strip().lower()
            alias = self.COLUMN_ALIASES.get(key)
            if alias:
                normalized[alias] = index
        return normalized

    def validate_structure(self, file_path: str) -> Tuple[bool, Optional[str]]:
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            worksheet = workbook.active
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            normalized = self._normalize_header(header_row)
            workbook.close()

            missing = [col for col in self.EXPECTED_COLUMNS if col in self.validator.REQUIRED_COLUMNS and col not in normalized]
            if missing:
                return False, f"Faltan columnas obligatorias: {', '.join(missing)}"
            return True, None
        except Exception as exc:
            return False, f"Error al leer el archivo Excel: {str(exc)}"

    def preview_rows(self, file_path: str, max_rows: int = 15) -> List[PreviewRow]:
        rows: List[PreviewRow] = []
        workbook = None
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            worksheet = workbook.active
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            normalized = self._normalize_header(header_row)

            for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if len(rows) >= max_rows:
                    break
                row_data = self._build_row_data(normalized, row)
                if not self._row_has_values(row_data):
                    continue
                valid, message = self.validator.validate_row_data(row_data, index)
                rows.append(PreviewRow(row_number=index, raw_data=row_data, validation_message=message))
        finally:
            if workbook is not None:
                workbook.close()
        return rows

    def iter_rows(self, file_path: str) -> Iterator[Tuple[int, Dict[str, Any]]]:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        normalized = self._normalize_header(header_row)

        try:
            for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = self._build_row_data(normalized, row)
                if not self._row_has_values(row_data):
                    continue
                yield row_index, row_data
        finally:
            workbook.close()

    def _build_row_data(self, normalized_header: Dict[str, int], row: Tuple[Any, ...]) -> Dict[str, Any]:
        row_data: Dict[str, Any] = {}
        for column_name, idx in normalized_header.items():
            if idx <= len(row):
                row_data[column_name] = row[idx - 1]
        return row_data

    def _row_has_values(self, row_data: Dict[str, Any]) -> bool:
        return any(value is not None and str(value).strip() for value in row_data.values())

    def generate_template(self, output_path: str) -> bool:
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Importación de Estudiantes'

            headers = [
                'Nombre',
                'Apellido',
                'Curso',
                'Aula',
                'Turno',
                'Tutor',
                'Matrícula',
                'Género',
                'FechaNacimiento',
                'Teléfono',
                'Año Escolar'
            ]

            for col_index, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=col_index, value=header)

            sample_rows = [
                ['Juan', 'Pérez', 'Primero', 'A', 'Mañana', 'María Pérez', '2024-001', 'M', '2009-03-12', '987654321', 2024],
                ['Ana', 'García', 'Segundo', 'B', 'Tarde', 'Carlos García', '2024-002', 'F', '2008-07-22', '956123478', 2024],
            ]
            for row_index, sample in enumerate(sample_rows, start=2):
                for col_index, value in enumerate(sample, start=1):
                    worksheet.cell(row=row_index, column=col_index, value=value)

            instructions = [
                'Instrucciones:',
                '1. No cambie los encabezados de columna.',
                '2. Use una sola hoja con datos de estudiantes.',
                '3. Curso, Aula y Turno deben corresponder a la configuración del sistema.',
                '4. Año Escolar debe ser numérico.',
                '5. FechaNacimiento puede ser fecha o texto YYYY-MM-DD.',
                '6. Matrícula es opcional pero recomendable para evitar duplicados.',
            ]
            sheet = workbook.create_sheet('Instrucciones')
            for row_index, text in enumerate(instructions, start=1):
                sheet.cell(row=row_index, column=1, value=text)

            workbook.save(output_path)
            return True
        except Exception:
            return False
